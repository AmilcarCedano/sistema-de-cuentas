"""
╔══════════════════════════════════════════════════════════════════════════════╗
║          GENERADOR DE REPORTE EXCEL – FINCONTROL                           ║
║──────────────────────────────────────────────────────────────────────────────║
║  Qué hace este script:                                                      ║
║   1. Lee un archivo .xlsx exportado desde FinControl                        ║
║   2. Extrae INGRESOS y EGRESOS por categoría                                ║
║   3. Genera un Excel bonito con:                                             ║
║      - Resumen de ingresos y egresos lado a lado con indicadores de barra   ║
║      - Balance final (Ingresos - Egresos = Neto)                            ║
║      - Detalle de cada transacción agrupada por categoría                   ║
║                                                                              ║
║  Instalación de dependencias (solo la primera vez):                         ║
║   pip install pandas openpyxl                                               ║
║                                                                              ║
║  Uso:                                                                        ║
║   python generar_reporte_fincontrol.py                                      ║
║   → Edita las variables ARCHIVO_ENTRADA y ARCHIVO_SALIDA abajo             ║
╚══════════════════════════════════════════════════════════════════════════════╝
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════════════════════════
# ► CONFIGURACIÓN – EDITA ESTO SEGÚN TU ARCHIVO
# ══════════════════════════════════════════════════════════════════════════════

ARCHIVO_ENTRADA = "Reporte_FinControl.xlsx"   # ← tu archivo exportado de FinControl
ARCHIVO_SALIDA  = "reporte_completo.xlsx"     # ← nombre del Excel que se va a generar

# Filas donde están los datos (basado en la estructura de FinControl)
# Ingresos: desde fila 5 hasta fila 46 (índices pandas: 4 a 45)
# Egresos:  desde fila 52 hasta fila 147 (índices pandas: 51 a 146)
# Si tu archivo tiene más o menos filas, ajusta estos valores
FILA_INICIO_INGRESOS = 4    # índice pandas (0-based), fila Excel 5
FILA_FIN_INGRESOS    = 46   # índice pandas exclusivo
FILA_INICIO_EGRESOS  = 51   # índice pandas (0-based), fila Excel 52
FILA_FIN_EGRESOS     = 147  # índice pandas exclusivo

# ══════════════════════════════════════════════════════════════════════════════
# ► PALETA DE COLORES POR CATEGORÍA
# Formato: "NOMBRE_CATEGORIA": ("COLOR_OSCURO", "COLOR_CLARO")
# COLOR_OSCURO → cabeceras  |  COLOR_CLARO → filas de datos
# Puedes agregar más categorías aquí con sus colores
# ══════════════════════════════════════════════════════════════════════════════

COLORES_EGRESOS = {
    "Centro_Medico":              ("1A5276", "D4E6F1"),  # azul oscuro / azul claro
    "GASTOS_CASA":                ("145A32", "D5F5E3"),  # verde oscuro / verde claro
    "Gastos de jercy ":           ("6E2F8E", "E8DAEF"),  # morado / lila
    "Gastos y ingresos de mamá ": ("7D6608", "FEF9E7"),  # dorado / amarillo claro
    "Gastos_Amilcar":             ("784212", "FDEBD0"),  # café / naranja claro
    "MEDICAMENTOS":               ("922B21", "FADBD8"),  # rojo oscuro / rosado
    "PASAJES":                    ("1F618D", "D6EAF8"),  # azul / celeste
    "Sin categoría":              ("424949", "EAECEE"),  # gris oscuro / gris claro
}

COLORES_INGRESOS = {
    "Centro_Medico":                ("117A65", "D1F2EB"),  # verde azulado
    "DEPOSITO_DE_CENTRO_HURANCHAL": ("1D8348", "D5F5E3"),  # verde
    "Sin categoría":                ("515A5A", "EAECEE"),  # gris
}

# Color por defecto si una categoría no está en el diccionario
COLOR_DEFAULT_OSCURO = "424949"
COLOR_DEFAULT_CLARO  = "EAECEE"


# ══════════════════════════════════════════════════════════════════════════════
# FUNCIONES AUXILIARES DE ESTILO
# ══════════════════════════════════════════════════════════════════════════════

def borde_fino(color="BDC3C7"):
    """Borde fino de 1px para celdas de datos."""
    lado = Side(style="thin", color=color)
    return Border(left=lado, right=lado, top=lado, bottom=lado)

def borde_medio(color="1B2631"):
    """Borde medio para filas de totales y cabeceras importantes."""
    lado = Side(style="medium", color=color)
    return Border(left=lado, right=lado, top=lado, bottom=lado)

def barra_indicador(pct, chars=18):
    """
    Crea una barra de progreso con bloques Unicode.
    Ejemplo: 65% → '████████████░░░░░░'
    """
    llenos = round(pct / 100 * chars)
    return "█" * llenos + "░" * (chars - llenos)


# ══════════════════════════════════════════════════════════════════════════════
# LECTURA Y LIMPIEZA DE DATOS
# ══════════════════════════════════════════════════════════════════════════════

def leer_datos(archivo):
    """
    Lee el archivo Excel de FinControl y retorna dos DataFrames:
    uno para ingresos y otro para egresos, ya limpios.
    """
    raw = pd.read_excel(archivo, sheet_name=0, header=None)
    columnas = ["Fecha", "Hora", "Categoria", "Concepto", "Nota", "Monto"]

    def limpiar(df_raw):
        df = df_raw.copy()
        df.columns = columnas
        df = df[df["Monto"].notna() & df["Categoria"].notna()]
        df["Monto"] = pd.to_numeric(df["Monto"], errors="coerce")
        df = df[df["Monto"].notna()]
        df["Fecha"] = pd.to_datetime(df["Fecha"], dayfirst=True, errors="coerce")
        return df

    egresos  = limpiar(raw.iloc[FILA_INICIO_EGRESOS:FILA_FIN_EGRESOS, :])
    ingresos = limpiar(raw.iloc[FILA_INICIO_INGRESOS:FILA_FIN_INGRESOS, :])
    return ingresos, egresos


# ══════════════════════════════════════════════════════════════════════════════
# BLOQUES DE ESCRITURA EN EXCEL
# ══════════════════════════════════════════════════════════════════════════════

def escribir_titulo_principal(ws):
    """Título grande que ocupa toda la fila superior."""
    ws.merge_cells("A1:O1")
    ws["A1"] = "REPORTE FINANCIERO COMPLETO – FINCONTROL"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1B2631")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:O2")
    ws["A2"] = "Ingresos y Egresos con detalle completo de transacciones por categoría"
    ws["A2"].font = Font(name="Arial", italic=True, size=9, color="FFFFFF")
    ws["A2"].fill = PatternFill("solid", fgColor="2E4057")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 8  # espaciador


def escribir_cabeceras_seccion(ws, fila):
    """
    Escribe los títulos 'EGRESOS – RESUMEN' e 'INGRESOS – RESUMEN'
    en las columnas A y I respectivamente.
    """
    ws.merge_cells(f"A{fila}:G{fila}")
    ws[f"A{fila}"] = "  📤  EGRESOS – RESUMEN"
    ws[f"A{fila}"].font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    ws[f"A{fila}"].fill = PatternFill("solid", fgColor="922B21")
    ws[f"A{fila}"].alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells(f"I{fila}:O{fila}")
    ws[f"I{fila}"] = "  📥  INGRESOS – RESUMEN"
    ws[f"I{fila}"].font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    ws[f"I{fila}"].fill = PatternFill("solid", fgColor="1D8348")
    ws[f"I{fila}"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[fila].height = 24


def escribir_fila_headers_resumen(ws, fila, col_inicio):
    """
    Cabecera de columnas para la tabla de resumen:
    Categoría | Total (S/) | N° Mov. | % Total | Indicador
    """
    headers = ["Categoría", "Total (S/)", "N° Mov.", "% Total", "Indicador", "", ""]
    for i, h in enumerate(headers):
        c = ws.cell(row=fila, column=col_inicio + i, value=h)
        c.font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1B2631")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = borde_fino("555555")
    ws.row_dimensions[fila].height = 20


def escribir_filas_resumen(ws, categorias_ordenadas, total_general, col_inicio, fila_inicio, mapa_colores):
    """
    Escribe una fila por cada categoría en la tabla de resumen.
    Retorna la siguiente fila disponible después de escribir todas.
    """
    fila = fila_inicio
    for cat, grupo in categorias_ordenadas:
        oscuro, claro = mapa_colores.get(cat, (COLOR_DEFAULT_OSCURO, COLOR_DEFAULT_CLARO))
        total = grupo["Monto"].sum()
        cantidad = len(grupo)
        pct = total / total_general * 100
        indicador = barra_indicador(pct)

        valores = [cat.strip(), total, cantidad, pct / 100, indicador, "", ""]
        for i, v in enumerate(valores):
            col = col_inicio + i
            c = ws.cell(row=fila, column=col, value=v)
            c.fill = PatternFill("solid", fgColor=claro)
            c.border = borde_fino()
            c.alignment = Alignment(horizontal="center", vertical="center")
            if i == 0:   # nombre categoría
                c.font = Font(name="Arial", bold=True, size=9, color=oscuro)
                c.alignment = Alignment(horizontal="left", vertical="center")
            elif i == 4: # indicador de barra
                c.font = Font(name="Courier New", size=7, color=oscuro)
            else:
                c.font = Font(name="Arial", size=9, color="1B2631")

        ws.cell(row=fila, column=col_inicio + 1).number_format = "#,##0.00"
        ws.cell(row=fila, column=col_inicio + 3).number_format = "0.0%"
        ws.row_dimensions[fila].height = 18
        fila += 1
    return fila


def escribir_fila_total_resumen(ws, fila, col_inicio, total, cantidad, color_fondo):
    """Fila de TOTAL GENERAL al final de cada tabla de resumen."""
    valores = ["TOTAL GENERAL", total, cantidad, 1.0, "█" * 18, "", ""]
    for i, v in enumerate(valores):
        c = ws.cell(row=fila, column=col_inicio + i, value=v)
        c.font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=color_fondo)
        c.border = borde_medio(color_fondo)
        c.alignment = Alignment(horizontal="left" if i == 0 else "center", vertical="center")
    ws.cell(row=fila, column=col_inicio + 1).number_format = "#,##0.00"
    ws.cell(row=fila, column=col_inicio + 3).number_format = "0.0%"
    ws.cell(row=fila, column=col_inicio + 4).font = Font(name="Courier New", size=7, color="F39C12")
    ws.row_dimensions[fila].height = 22


def escribir_balance_neto(ws, fila, total_ingresos, total_egresos):
    """
    Escribe la pequeña tabla de balance:
      Total Ingresos
      Total Egresos
      Monto Neto (diferencia)
    """
    neto = total_ingresos - total_egresos

    # Título
    ws.merge_cells(f"A{fila}:O{fila}")
    ws[f"A{fila}"] = "  💰  BALANCE FINAL DE LA CUENTA"
    ws[f"A{fila}"].font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    ws[f"A{fila}"].fill = PatternFill("solid", fgColor="1B2631")
    ws[f"A{fila}"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[fila].height = 24
    fila += 1

    # Headers de la mini tabla
    for col, h in enumerate(["Concepto", "Monto (S/)"], 1):
        c = ws.cell(row=fila, column=col, value=h)
        c.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="2E4057")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = borde_fino()
    ws.row_dimensions[fila].height = 20
    fila += 1

    # Filas de ingresos y egresos
    items = [
        ("📥  Total Ingresos", total_ingresos, "1D8348", "D5F5E3"),
        ("📤  Total Egresos",  total_egresos,  "7B241C", "FADBD8"),
    ]
    for label, val, oscuro, claro in items:
        ws.cell(row=fila, column=1, value=label).font = Font(name="Arial", bold=True, size=10, color=oscuro)
        ws.cell(row=fila, column=1).fill = PatternFill("solid", fgColor=claro)
        ws.cell(row=fila, column=1).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=fila, column=1).border = borde_fino()
        c = ws.cell(row=fila, column=2, value=val)
        c.font = Font(name="Arial", bold=True, size=10, color=oscuro)
        c.fill = PatternFill("solid", fgColor=claro)
        c.number_format = "#,##0.00"
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = borde_fino()
        ws.row_dimensions[fila].height = 20
        fila += 1

    # Fila neto
    color_neto = "1D6A27" if neto >= 0 else "922B21"
    ws.cell(row=fila, column=1, value="✅  MONTO NETO (Ingresos − Egresos)").font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    ws.cell(row=fila, column=1).fill = PatternFill("solid", fgColor=color_neto)
    ws.cell(row=fila, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=fila, column=1).border = borde_medio(color_neto)
    c = ws.cell(row=fila, column=2, value=neto)
    c.font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=color_neto)
    c.number_format = "#,##0.00"
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = borde_medio(color_neto)
    ws.row_dimensions[fila].height = 26
    return fila + 2  # deja un espacio


def escribir_bloque_detalle(ws, cat, grupo, total_general, col_inicio, mapa_colores, fila_inicio):
    """
    Escribe el bloque de detalle de una categoría:
    - Cabecera con nombre, total, cantidad y % 
    - Fila de headers (#, Fecha, Concepto, Nota, Monto)
    - Una fila por cada transacción (alternando color)
    - Fila de SUBTOTAL al final
    Retorna la siguiente fila disponible.
    """
    oscuro, claro = mapa_colores.get(cat, (COLOR_DEFAULT_OSCURO, COLOR_DEFAULT_CLARO))
    total_cat = grupo["Monto"].sum()
    cnt = len(grupo)
    pct = total_cat / total_general * 100
    fila = fila_inicio

    col_inicio_letra = get_column_letter(col_inicio)
    col_fin_letra    = get_column_letter(col_inicio + 6)

    # Cabecera de categoría
    ws.merge_cells(f"{col_inicio_letra}{fila}:{col_fin_letra}{fila}")
    ws[f"{col_inicio_letra}{fila}"] = (
        f"  {cat.strip().upper()}   |   S/ {total_cat:,.2f}"
        f"   |   {cnt} mov.   |   {pct:.1f}%"
    )
    ws[f"{col_inicio_letra}{fila}"].font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
    ws[f"{col_inicio_letra}{fila}"].fill = PatternFill("solid", fgColor=oscuro)
    ws[f"{col_inicio_letra}{fila}"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[fila].height = 20
    fila += 1

    # Headers de columnas
    for i, h in enumerate(["#", "Fecha", "Concepto", "Nota", "Monto (S/)", "", ""]):
        c = ws.cell(row=fila, column=col_inicio + i, value=h)
        c.font = Font(name="Arial", bold=True, size=8, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=oscuro)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = borde_fino(oscuro)
    ws.row_dimensions[fila].height = 16
    fila += 1

    # Filas de transacciones
    for idx, (_, row) in enumerate(grupo.sort_values("Fecha").iterrows(), 1):
        color_fila = claro if idx % 2 == 0 else "FFFFFF"
        fecha_str = row["Fecha"].strftime("%d/%m/%Y") if pd.notna(row["Fecha"]) else ""
        nota = str(row["Nota"]).strip() if pd.notna(row["Nota"]) and str(row["Nota"]) != "nan" else ""
        valores = [idx, fecha_str, str(row["Concepto"]).strip(), nota, row["Monto"], "", ""]

        for i, v in enumerate(valores):
            col = col_inicio + i
            c = ws.cell(row=fila, column=col, value=v)
            c.font = Font(name="Arial", size=8)
            c.fill = PatternFill("solid", fgColor=color_fila)
            c.alignment = Alignment(
                horizontal="center" if i in [0, 1, 4] else "left",
                vertical="center"
            )
            c.border = borde_fino()
        ws.cell(row=fila, column=col_inicio + 4).number_format = "#,##0.00"
        ws.row_dimensions[fila].height = 16
        fila += 1

    # Subtotal
    col_d_letra = get_column_letter(col_inicio + 3)
    ws.merge_cells(f"{col_inicio_letra}{fila}:{col_d_letra}{fila}")
    ws[f"{col_inicio_letra}{fila}"] = "SUBTOTAL"
    ws[f"{col_inicio_letra}{fila}"].font = Font(name="Arial", bold=True, size=8, color="FFFFFF")
    ws[f"{col_inicio_letra}{fila}"].fill = PatternFill("solid", fgColor=oscuro)
    ws[f"{col_inicio_letra}{fila}"].alignment = Alignment(horizontal="right")
    ws[f"{col_inicio_letra}{fila}"].border = borde_medio(oscuro)

    c5 = ws.cell(row=fila, column=col_inicio + 4, value=total_cat)
    c5.font = Font(name="Arial", bold=True, size=8, color="FFFFFF")
    c5.fill = PatternFill("solid", fgColor=oscuro)
    c5.number_format = "#,##0.00"
    c5.alignment = Alignment(horizontal="center")
    c5.border = borde_medio(oscuro)

    for x in [5, 6]:
        cc = ws.cell(row=fila, column=col_inicio + x)
        cc.fill = PatternFill("solid", fgColor=oscuro)
        cc.border = borde_medio(oscuro)

    ws.row_dimensions[fila].height = 18
    fila += 1
    ws.row_dimensions[fila].height = 6  # mini espaciador
    fila += 1
    return fila


def configurar_anchos(ws):
    """Ajusta el ancho de cada columna."""
    # Columnas A-G (egresos): #, Fecha, Concepto, Nota, Monto, vacía, vacía
    anchos_eg = [4, 11, 22, 26, 12, 1, 1]
    for i, w in enumerate(anchos_eg):
        ws.column_dimensions[get_column_letter(1 + i)].width = w  # cols 1-7

    # Columna H (separador)
    ws.column_dimensions["H"].width = 2

    # Columnas I-O (ingresos): misma distribución
    for i, w in enumerate(anchos_eg):
        ws.column_dimensions[get_column_letter(9 + i)].width = w  # cols 9-15


# ══════════════════════════════════════════════════════════════════════════════
# FUNCIÓN PRINCIPAL
# ══════════════════════════════════════════════════════════════════════════════

def generar_reporte(archivo_entrada, archivo_salida):
    """
    Función principal que orquesta todo el proceso:
    1. Lee los datos
    2. Crea el workbook
    3. Escribe cada sección
    4. Guarda el archivo
    """
    print(f"Leyendo datos de: {archivo_entrada}")
    ingresos, egresos = leer_datos(archivo_entrada)

    total_ingresos = ingresos["Monto"].sum()
    total_egresos  = egresos["Monto"].sum()
    neto           = total_ingresos - total_egresos

    print(f"  Total ingresos : S/ {total_ingresos:,.2f}")
    print(f"  Total egresos  : S/ {total_egresos:,.2f}")
    print(f"  Neto           : S/ {neto:,.2f}")

    # Ordenar categorías de mayor a menor gasto/ingreso
    eg_cats = sorted(egresos.groupby("Categoria"),  key=lambda x: -x[1]["Monto"].sum())
    in_cats = sorted(ingresos.groupby("Categoria"), key=lambda x: -x[1]["Monto"].sum())

    # Columnas de inicio
    COL_EGRESOS  = 1   # columna A
    COL_SPACER   = 8   # columna H (separador visual)
    COL_INGRESOS = 9   # columna I

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Completo"

    # ── 1. Título principal ───────────────────────────────────────────────────
    escribir_titulo_principal(ws)

    # ── 2. Cabeceras de sección (fila 4) ─────────────────────────────────────
    escribir_cabeceras_seccion(ws, fila=4)

    # ── 3. Headers de resumen (fila 5) ────────────────────────────────────────
    escribir_fila_headers_resumen(ws, fila=5, col_inicio=COL_EGRESOS)
    escribir_fila_headers_resumen(ws, fila=5, col_inicio=COL_INGRESOS)

    # ── 4. Filas de resumen por categoría (desde fila 6) ─────────────────────
    eg_res_end = escribir_filas_resumen(ws, eg_cats, total_egresos,  COL_EGRESOS,  6, COLORES_EGRESOS)
    in_res_end = escribir_filas_resumen(ws, in_cats, total_ingresos, COL_INGRESOS, 6, COLORES_INGRESOS)
    fila_totales = max(eg_res_end, in_res_end)

    # ── 5. Fila de totales generales ──────────────────────────────────────────
    escribir_fila_total_resumen(ws, fila_totales, COL_EGRESOS,  total_egresos,  len(egresos),  "7B241C")
    escribir_fila_total_resumen(ws, fila_totales, COL_INGRESOS, total_ingresos, len(ingresos), "1D8348")

    # ── 6. Balance / Neto ─────────────────────────────────────────────────────
    fila_actual = fila_totales + 2
    fila_actual = escribir_balance_neto(ws, fila_actual, total_ingresos, total_egresos)

    # ── 7. Título de sección de detalle ───────────────────────────────────────
    ws.merge_cells(f"A{fila_actual}:O{fila_actual}")
    ws[f"A{fila_actual}"] = "  📋  DETALLE COMPLETO DE TRANSACCIONES POR CATEGORÍA"
    ws[f"A{fila_actual}"].font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    ws[f"A{fila_actual}"].fill = PatternFill("solid", fgColor="2E4057")
    ws[f"A{fila_actual}"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[fila_actual].height = 24
    fila_actual += 1

    # Sub-etiquetas EGRESOS | INGRESOS
    ws.merge_cells(f"A{fila_actual}:G{fila_actual}")
    ws[f"A{fila_actual}"] = "  📤  EGRESOS"
    ws[f"A{fila_actual}"].font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    ws[f"A{fila_actual}"].fill = PatternFill("solid", fgColor="7B241C")
    ws[f"A{fila_actual}"].alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells(f"I{fila_actual}:O{fila_actual}")
    ws[f"I{fila_actual}"] = "  📥  INGRESOS"
    ws[f"I{fila_actual}"].font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    ws[f"I{fila_actual}"].fill = PatternFill("solid", fgColor="1D8348")
    ws[f"I{fila_actual}"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[fila_actual].height = 20
    fila_actual += 1

    # ── 8. Bloques de detalle por categoría (lado a lado) ─────────────────────
    fila_eg = fila_actual
    fila_in = fila_actual

    for cat, grupo in eg_cats:
        fila_eg = escribir_bloque_detalle(ws, cat, grupo, total_egresos, COL_EGRESOS, COLORES_EGRESOS, fila_eg)

    for cat, grupo in in_cats:
        fila_in = escribir_bloque_detalle(ws, cat, grupo, total_ingresos, COL_INGRESOS, COLORES_INGRESOS, fila_in)

    # ── 9. Anchos de columna ──────────────────────────────────────────────────
    configurar_anchos(ws)

    # ── 10. Guardar ───────────────────────────────────────────────────────────
    wb.save(archivo_salida)
    print(f"\n✅ Reporte generado: {archivo_salida}")


# ══════════════════════════════════════════════════════════════════════════════
# PUNTO DE ENTRADA
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    generar_reporte(ARCHIVO_ENTRADA, ARCHIVO_SALIDA)
